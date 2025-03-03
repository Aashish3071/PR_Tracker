from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, FileResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import AnonymousUser
import csv
import pandas as pd
from .forms import CoverageForm, CustomFormulaForm, ReportTemplateForm, GenerateReportForm
from .models import Coverage, Publication, Edition, Rate, Client, Campaign, CustomFormula, ReportTemplate, GeneratedReport
from django.db.models import Sum, Count, Avg, F
from django.db.models.functions import Coalesce
from decimal import Decimal
import openpyxl
from datetime import datetime
from django.db import models
from django.db.models import Q
import os
from django.conf import settings
from django.urls import reverse

@login_required(login_url='/admin/login/')
def home(request):
    # Get statistics for dashboard
    total_coverages = Coverage.objects.count()
    print_coverages = Coverage.objects.filter(type='Print').count()
    online_coverages = Coverage.objects.filter(type='Online').count()
    
    # Calculate percentages
    print_percentage = round((print_coverages / total_coverages) * 100) if total_coverages > 0 else 0
    online_percentage = round((online_coverages / total_coverages) * 100) if total_coverages > 0 else 0
    
    # Calculate total AVE
    total_ave = Coverage.objects.aggregate(total=Sum('ave'))['total'] or 0
    
    # Get recent coverages
    recent_coverages = Coverage.objects.select_related('publication', 'edition').order_by('-date')[:10]
    
    # Get top publications by coverage count and AVE
    top_publications = Coverage.objects.values('publication__name').annotate(
        count=Count('id'),
        total_ave=Sum('ave')
    ).order_by('-count')[:5]
    
    # Get all publications and editions for filters
    all_publications = Publication.objects.all()
    all_editions = Edition.objects.select_related('publication').all()
    
    context = {
        'total_coverages': total_coverages,
        'print_coverages': print_coverages,
        'online_coverages': online_coverages,
        'print_percentage': print_percentage,
        'online_percentage': online_percentage,
        'total_ave': total_ave,
        'recent_coverages': recent_coverages,
        'top_publications': top_publications,
        'all_publications': all_publications,
        'all_editions': all_editions,
    }
    
    return render(request, 'core/dashboard.html', context)

@login_required(login_url='/admin/login/')
def add_coverage(request):
    if request.method == 'POST':
        form = CoverageForm(request.POST)
        if form.is_valid():
            try:
                coverage = form.save(commit=False)
                coverage.user = request.user
                
                # Ensure required fields are set
                if not coverage.client:
                    form.add_error('client', 'Client is required')
                    return render(request, 'core/add_coverage.html', {'form': form})
                    
                if not coverage.publication:
                    form.add_error('publication', 'Publication is required')
                    return render(request, 'core/add_coverage.html', {'form': form})
                    
                if not coverage.edition:
                    form.add_error('edition', 'Edition is required')
                    return render(request, 'core/add_coverage.html', {'form': form})
                
                # Handle online coverage
                if coverage.type == 'Online' or coverage.is_online:
                    coverage.is_online = True
                    if not coverage.ave:
                        form.add_error('ave', 'AVE is required for online coverage')
                        return render(request, 'core/add_coverage.html', {'form': form})
                
                # Save the coverage
                coverage.save()
                messages.success(request, "Coverage added successfully.")
                return redirect('coverage_list')
            except Exception as e:
                messages.error(request, f"Error saving coverage: {str(e)}")
        else:
            # Form is invalid, display errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Pre-fill form with query parameters if provided
        initial_data = {}
        if request.GET.get('client'):
            try:
                client_id = int(request.GET.get('client'))
                initial_data['client'] = client_id
                
                # If campaign is also provided, add it
                if request.GET.get('campaign'):
                    try:
                        campaign_id = int(request.GET.get('campaign'))
                        initial_data['campaign'] = campaign_id
                    except ValueError:
                        pass
            except ValueError:
                pass
                
        if request.GET.get('publication'):
            try:
                publication_id = int(request.GET.get('publication'))
                initial_data['publication'] = publication_id
                
                # If edition is also provided, add it
                if request.GET.get('edition'):
                    try:
                        edition_id = int(request.GET.get('edition'))
                        initial_data['edition'] = edition_id
                    except ValueError:
                        pass
            except ValueError:
                pass
        
        form = CoverageForm(initial=initial_data)
    
    return render(request, 'core/add_coverage.html', {'form': form})

@login_required(login_url='/admin/login/')
def coverage_list(request):
    coverages = Coverage.objects.filter(user=request.user)
    print_count = coverages.filter(is_online=False).count()
    online_count = coverages.filter(is_online=True).count()
    return render(request, 'core/coverage_list.html', {
        'coverages': coverages,
        'print_count': print_count,
        'online_count': online_count
    })

@login_required(login_url='/admin/login/')
def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Check if file is an Excel file
        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, 'Please upload a valid Excel file (.xls or .xlsx)')
            return redirect('upload_excel')
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            # Get headers
            headers = [cell.value for cell in sheet[1]]
            
            # Required columns
            required_columns = ['Date', 'Publication', 'Edition', 'Headline', 'Type', 'Size']
            
            # Check if all required columns exist
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
                return redirect('upload_excel')
            
            # Get column indices
            date_idx = headers.index('Date')
            publication_idx = headers.index('Publication')
            edition_idx = headers.index('Edition')
            headline_idx = headers.index('Headline')
            type_idx = headers.index('Type')
            size_idx = headers.index('Size')
            
            # Optional columns
            page_idx = headers.index('Page') if 'Page' in headers else None
            position_idx = headers.index('Position') if 'Position' in headers else None
            client_idx = headers.index('Client') if 'Client' in headers else None
            campaign_idx = headers.index('Campaign') if 'Campaign' in headers else None
            
            # Process rows
            coverages_created = 0
            errors = []
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    # Get values
                    date_value = row[date_idx]
                    publication_name = row[publication_idx]
                    edition_name = row[edition_idx]
                    headline = row[headline_idx]
                    coverage_type = row[type_idx]
                    size = row[size_idx]
                    
                    # Optional values
                    page = row[page_idx] if page_idx is not None else None
                    position = row[position_idx] if position_idx is not None else None
                    client_name = row[client_idx] if client_idx is not None else None
                    campaign_name = row[campaign_idx] if campaign_idx is not None else None
                    
                    # Validate date
                    if isinstance(date_value, str):
                        try:
                            date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
                        except ValueError:
                            try:
                                date_value = datetime.strptime(date_value, '%d/%m/%Y').date()
                            except ValueError:
                                errors.append(f"Row {row_idx}: Invalid date format. Use YYYY-MM-DD or DD/MM/YYYY.")
                                continue
                    
                    # Get or create publication
                    publication, _ = Publication.objects.get_or_create(name=publication_name)
                    
                    # Get or create edition
                    edition, _ = Edition.objects.get_or_create(
                        name=edition_name,
                        publication=publication
                    )
                    
                    # Get or create client if provided
                    client = None
                    if client_name:
                        client, _ = Client.objects.get_or_create(name=client_name)
                    
                    # Get or create campaign if provided
                    campaign = None
                    if campaign_name and client:
                        campaign, _ = Campaign.objects.get_or_create(
                            name=campaign_name,
                            client=client
                        )
                    
                    # Create coverage
                    coverage = Coverage(
                        date=date_value,
                        publication=publication,
                        edition=edition,
                        headline=headline,
                        type=coverage_type,
                        size=size,
                        page=page,
                        position=position,
                    client=client,
                    campaign=campaign,
                        user=request.user
                    )
                    
                    # Calculate AVE
                    coverage.calculate_ave()
                    coverage.save()
                    
                    coverages_created += 1
                
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            # Show success message
            if coverages_created > 0:
                messages.success(request, f"Successfully imported {coverages_created} coverages.")
            
            # Show errors if any
            if errors:
                messages.warning(request, f"Encountered {len(errors)} errors during import.")
                for error in errors[:5]:  # Show first 5 errors
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.error(request, f"... and {len(errors) - 5} more errors.")
            
            return redirect('coverage_list')
        
        except Exception as e:
            messages.error(request, f"Error processing Excel file: {str(e)}")
            return redirect('upload_excel')

    return render(request, 'core/upload_excel.html')

@login_required(login_url='/admin/login/')
def coverage_report(request):
    coverages = Coverage.objects.filter(user=request.user)
    if 'start_date' in request.GET:
        coverages = coverages.filter(date__gte=request.GET['start_date'])
    if 'end_date' in request.GET:
        coverages = coverages.filter(date__lte=request.GET['end_date'])

    print_coverages = coverages.filter(type='Print')
    online_coverages = coverages.filter(type='Online')

    if 'export' in request.GET and request.GET['export'] == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="coverage_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Sr. No.', 'Date', 'Publication', 'Edition', 'Headline', 'Page', 'Size', 'Rate', 'AVE'])
        for i, cov in enumerate(print_coverages, 1):
            writer.writerow([
                i, cov.date, cov.publication.name, cov.edition.name, cov.headline,
                cov.page, cov.size, cov.rate_per_sq_cm, cov.ave
            ])
        writer.writerow([])
        writer.writerow(['Online Coverages'])
        writer.writerow(['Sr. No.', 'Date', 'Publication', 'Headline', 'AVE'])
        for i, cov in enumerate(online_coverages, 1):
            writer.writerow([i, cov.date, cov.publication.name, cov.headline, cov.ave])
        return response

    return render(request, 'core/report.html', {
        'print_coverages': print_coverages,
        'online_coverages': online_coverages,
    })

def ave_report(request):
    # Get all coverages
    coverages = Coverage.objects.select_related('publication', 'edition').all()
    
    # Separate print and online coverages
    print_coverages = coverages.filter(type='Print').order_by('date')
    online_coverages = coverages.filter(type='Online').order_by('date')
    
    # Calculate print statistics
    print_count = print_coverages.count()
    print_total_size = print_coverages.aggregate(total=Sum('size'))['total'] or 0
    print_ave = print_coverages.aggregate(total=Sum('ave'))['total'] or 0
    
    # Calculate online statistics
    online_count = online_coverages.count()
    online_ave = online_coverages.aggregate(total=Sum('ave'))['total'] or 0
    
    # Calculate total statistics
    total_count = print_count + online_count
    total_ave = print_ave + online_ave
    
    # Calculate percentages
    print_percentage = (print_count / total_count * 100) if total_count > 0 else 0
    online_percentage = (online_count / total_count * 100) if total_count > 0 else 0
    print_ave_percentage = (print_ave / total_ave * 100) if total_ave > 0 else 0
    online_ave_percentage = (online_ave / total_ave * 100) if total_ave > 0 else 0
    
    # Get top publications by AVE
    top_publications = Coverage.objects.values('publication__name').annotate(
        count=Count('id'),
        total_ave=Sum('ave')
    ).order_by('-total_ave')[:10]
    
    context = {
        'print_coverages': print_coverages,
        'online_coverages': online_coverages,
        'print_count': print_count,
        'online_count': online_count,
        'total_count': total_count,
        'print_total_size': print_total_size,
        'print_ave': print_ave,
        'online_ave': online_ave,
        'total_ave': total_ave,
        'print_percentage': print_percentage,
        'online_percentage': online_percentage,
        'print_ave_percentage': print_ave_percentage,
        'online_ave_percentage': online_ave_percentage,
        'top_publications': top_publications,
    }
    
    return render(request, 'core/ave_report.html', context)

def export_ave_report(request):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AVE Report"
    
    # Add title and date
    ws['A1'] = "Weekly Coverage Sheet"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    ws['A2'] = f"Generated on: {datetime.now().strftime('%d %B %Y')}"
    
    # Add headers for print section
    ws['A4'] = "PRINT"
    ws['A4'].font = openpyxl.styles.Font(bold=True)
    
    headers = ['Sr. No.', 'Date', 'Publication', 'Editions', 'Headline', 'Page', 'Size (sq cms)', 'Rate', 'AVE']
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws[f'{col_letter}5'] = header
        ws[f'{col_letter}5'].font = openpyxl.styles.Font(bold=True)
    
    # Add print coverage data
    print_coverages = Coverage.objects.filter(type='Print').select_related('publication', 'edition').order_by('date')
    row_num = 6
    for i, coverage in enumerate(print_coverages, 1):
        ws[f'A{row_num}'] = i
        ws[f'B{row_num}'] = coverage.date.strftime('%d-%b-%y')
        ws[f'C{row_num}'] = coverage.publication.name
        ws[f'D{row_num}'] = coverage.edition.name
        ws[f'E{row_num}'] = coverage.headline
        ws[f'F{row_num}'] = coverage.page
        ws[f'G{row_num}'] = coverage.size
        ws[f'H{row_num}'] = coverage.rate_per_sq_cm
        ws[f'I{row_num}'] = coverage.ave
        row_num += 1
    
    # Add print total row
    print_total_row = row_num
    ws[f'F{print_total_row}'] = "Total"
    ws[f'F{print_total_row}'].font = openpyxl.styles.Font(bold=True)
    ws[f'I{print_total_row}'] = f"=SUM(I6:I{row_num-1})"
    ws[f'I{print_total_row}'].font = openpyxl.styles.Font(bold=True)
    
    # Add headers for online section
    row_num += 2
    ws[f'A{row_num}'] = "ONLINE"
    ws[f'A{row_num}'].font = openpyxl.styles.Font(bold=True)
    
    online_headers = ['Sr. No.', 'Date', 'Publication', 'Editions', 'Headline', 'AVE']
    row_num += 1
    for col_num, header in enumerate(online_headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws[f'{col_letter}{row_num}'] = header
        ws[f'{col_letter}{row_num}'].font = openpyxl.styles.Font(bold=True)
    
    # Add online coverage data
    online_coverages = Coverage.objects.filter(type='Online').select_related('publication', 'edition').order_by('date')
    row_num += 1
    online_start_row = row_num
    for i, coverage in enumerate(online_coverages, 1):
        ws[f'A{row_num}'] = i
        ws[f'B{row_num}'] = coverage.date.strftime('%d-%b-%y')
        ws[f'C{row_num}'] = coverage.publication.name
        ws[f'D{row_num}'] = coverage.edition.name
        ws[f'E{row_num}'] = coverage.headline
        ws[f'F{row_num}'] = coverage.ave
        row_num += 1
    
    # Add online total row
    online_total_row = row_num
    ws[f'E{online_total_row}'] = "Total"
    ws[f'E{online_total_row}'].font = openpyxl.styles.Font(bold=True)
    ws[f'F{online_total_row}'] = f"=SUM(F{online_start_row}:F{row_num-1})"
    ws[f'F{online_total_row}'].font = openpyxl.styles.Font(bold=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=AVE_Report.xlsx'
    wb.save(response)
    
    return response

@login_required(login_url='/admin/login/')
def custom_formula_list(request):
    formulas = CustomFormula.objects.filter(user=request.user)
    return render(request, 'core/custom_formula_list.html', {'formulas': formulas})

@login_required(login_url='/admin/login/')
def add_custom_formula(request):
    if request.method == 'POST':
        form = CustomFormulaForm(request.POST)
        if form.is_valid():
            formula = form.save(commit=False)
            formula.user = request.user
            formula.save()
            messages.success(request, "Custom formula added successfully.")
            return redirect('custom_formula_list')
    else:
        form = CustomFormulaForm()
    
    # Get available variables for the template
    available_variables = CustomFormula.AVAILABLE_VARIABLES
    
    return render(request, 'core/add_custom_formula.html', {
        'form': form,
        'available_variables': available_variables
    })

@login_required(login_url='/admin/login/')
def edit_custom_formula(request, formula_id):
    formula = get_object_or_404(CustomFormula, id=formula_id, user=request.user)
    
    if request.method == 'POST':
        form = CustomFormulaForm(request.POST, instance=formula)
        if form.is_valid():
            form.save()
            messages.success(request, "Custom formula updated successfully.")
            return redirect('custom_formula_list')
    else:
        form = CustomFormulaForm(instance=formula)
    
    # Get available variables for the template
    available_variables = CustomFormula.AVAILABLE_VARIABLES
    
    return render(request, 'core/edit_custom_formula.html', {
        'form': form,
        'formula': formula,
        'available_variables': available_variables
    })

@login_required(login_url='/admin/login/')
def delete_custom_formula(request, formula_id):
    formula = get_object_or_404(CustomFormula, id=formula_id, user=request.user)
    
    if request.method == 'POST':
        formula.delete()
        messages.success(request, "Custom formula deleted successfully.")
        return redirect('custom_formula_list')
    
    return render(request, 'core/delete_custom_formula.html', {'formula': formula})

@login_required(login_url='/admin/login/')
def test_custom_formula(request, formula_id):
    formula = get_object_or_404(CustomFormula, id=formula_id, user=request.user)
    
    if request.method == 'POST':
        # Get test values from the form
        test_data = {}
        for var, _ in CustomFormula.AVAILABLE_VARIABLES:
            if var in request.POST:
                try:
                    test_data[var] = float(request.POST.get(var, 0))
                except ValueError:
                    test_data[var] = 0
        
        # Calculate the result
        result = formula.calculate(test_data)
        
        return render(request, 'core/test_custom_formula.html', {
            'formula': formula,
            'test_data': test_data,
            'result': result,
            'available_variables': CustomFormula.AVAILABLE_VARIABLES
        })
    
    # Initialize with default values
    default_values = {var: 100 for var, _ in CustomFormula.AVAILABLE_VARIABLES}
    
    return render(request, 'core/test_custom_formula.html', {
        'formula': formula,
        'test_data': default_values,
        'available_variables': CustomFormula.AVAILABLE_VARIABLES
    })

@login_required(login_url='/admin/login/')
def search(request):
    query = request.GET.get('q', '')
    publication_id = request.GET.get('publication', '')
    edition_id = request.GET.get('edition', '')
    coverage_type = request.GET.get('type', '')
    
    # Always get all publications, editions, and clients for dropdowns
    all_publications = Publication.objects.all().order_by('name')
    all_editions = Edition.objects.select_related('publication').all().order_by('publication__name', 'name')
    all_clients = Client.objects.all().order_by('name')
    all_campaigns = Campaign.objects.select_related('client').all().order_by('client__name', 'name')
    
    # Initialize selected items
    selected_publication = None
    selected_edition = None
    selected_type = coverage_type
    
    # Base queryset for coverages
    coverages_queryset = Coverage.objects.select_related('publication', 'edition', 'client', 'campaign')
    
    # Apply filters if provided
    if publication_id:
        try:
            selected_publication = Publication.objects.get(id=publication_id)
            coverages_queryset = coverages_queryset.filter(publication_id=publication_id)
        except Publication.DoesNotExist:
            pass
    
    if edition_id:
        try:
            selected_edition = Edition.objects.get(id=edition_id)
            coverages_queryset = coverages_queryset.filter(edition_id=edition_id)
        except Edition.DoesNotExist:
            pass
    
    if coverage_type:
        coverages_queryset = coverages_queryset.filter(type=coverage_type)
    
    # Initialize querysets
    coverages = Coverage.objects.none()
    publications = Publication.objects.none()
    editions = Edition.objects.none()
    clients = Client.objects.none()
    campaigns = Campaign.objects.none()
    formulas = CustomFormula.objects.none()
    
    # If query is provided, search across models
    if query:
        coverages = coverages_queryset.filter(
            Q(headline__icontains=query) |
            Q(publication__name__icontains=query) |
            Q(edition__name__icontains=query) |
            Q(client__name__icontains=query) |
            Q(campaign__name__icontains=query)
        ).distinct().order_by('-date')
        
        publications = Publication.objects.filter(name__icontains=query).order_by('name')
        editions = Edition.objects.select_related('publication').filter(
            Q(name__icontains=query) | Q(publication__name__icontains=query)
        ).order_by('publication__name', 'name')
        clients = Client.objects.filter(name__icontains=query).order_by('name')
        campaigns = Campaign.objects.select_related('client').filter(
            Q(name__icontains=query) | Q(client__name__icontains=query)
        ).order_by('client__name', 'name')
        formulas = CustomFormula.objects.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        ).order_by('name')
    else:
        # If no query but filters are applied, just filter coverages
        coverages = coverages_queryset.order_by('-date')
        
        # Load some default data for the tabs
        publications = all_publications[:20]
        editions = all_editions[:20]
        clients = all_clients[:20]
        campaigns = all_campaigns[:20]
        formulas = CustomFormula.objects.all().order_by('name')[:20]
    
    # Check if we have any results for each category
    has_publications = publications.exists()
    has_editions = editions.exists()
    has_clients = clients.exists()
    has_campaigns = campaigns.exists()
    has_formulas = formulas.exists()
    
    total_results = (
        coverages.count() +
        publications.count() +
        editions.count() +
        clients.count() +
        campaigns.count() +
        formulas.count()
    )
    
    context = {
        'query': query,
        'coverages': coverages,
        'publications': publications,
        'editions': editions,
        'clients': clients,
        'campaigns': campaigns,
        'formulas': formulas,
        'total_results': total_results,
        'all_publications': all_publications,
        'all_editions': all_editions,
        'all_clients': all_clients,
        'all_campaigns': all_campaigns,
        'selected_publication': selected_publication,
        'selected_edition': selected_edition,
        'selected_type': selected_type,
        'has_publications': has_publications,
        'has_editions': has_editions,
        'has_clients': has_clients,
        'has_campaigns': has_campaigns,
        'has_formulas': has_formulas,
        'search_query': query,  # Add the search query for "Add New" buttons
    }
    
    return render(request, 'core/search.html', context)

# Report Template Views
@login_required(login_url='/admin/login/')
def report_template_list(request):
    templates = ReportTemplate.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'core/report_template_list.html', {
        'templates': templates
    })

@login_required(login_url='/admin/login/')
def upload_report_template(request):
    if request.method == 'POST':
        form = ReportTemplateForm(request.POST, request.FILES)
        if form.is_valid():
            template = form.save(commit=False)
            template.user = request.user
            template.save()
            
            # Redirect to the template mapping page
            messages.success(request, "Template uploaded successfully. Now map the fields.")
            return redirect('edit_report_template', template_id=template.id)
    else:
        form = ReportTemplateForm()
    
    return render(request, 'core/upload_report_template.html', {
        'form': form
    })

@login_required(login_url='/admin/login/')
def edit_report_template(request, template_id):
    template = get_object_or_404(ReportTemplate, id=template_id, user=request.user)
    
    # Get available fields from the database
    available_fields = {
        'coverage': [
            'date', 'headline', 'type', 'size', 'page', 'position', 'ave', 
            'publication__name', 'edition__name', 'client__name', 'campaign__name'
        ],
        'publication': ['name', 'rate_print', 'rate_online'],
        'edition': ['name'],
        'client': ['name'],
        'campaign': ['name', 'start_date', 'end_date']
    }
    
    if request.method == 'POST':
        # Save field mappings
        field_mappings = {}
        for key, value in request.POST.items():
            if key.startswith('mapping_'):
                excel_column = key.replace('mapping_', '')
                db_field = value
                if db_field:  # Only save non-empty mappings
                    field_mappings[excel_column] = db_field
        
        template.field_mappings = field_mappings
        template.save()
        messages.success(request, "Field mappings saved successfully.")
        return redirect('report_template_list')
    
    # Read Excel file to get column headers
    try:
        wb = openpyxl.load_workbook(template.template_file.path, read_only=True)
        sheet = wb.active
        excel_columns = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                excel_columns.append(cell_value)
    except Exception as e:
        messages.error(request, f"Error reading template file: {str(e)}")
        excel_columns = []
    
    return render(request, 'core/edit_report_template.html', {
        'template': template,
        'excel_columns': excel_columns,
        'available_fields': available_fields,
        'field_mappings': template.field_mappings
    })

@login_required(login_url='/admin/login/')
def delete_report_template(request, template_id):
    template = get_object_or_404(ReportTemplate, id=template_id, user=request.user)
    
    if request.method == 'POST':
        template.delete()
        messages.success(request, "Report template deleted successfully.")
        return redirect('report_template_list')
    
    return render(request, 'core/delete_report_template.html', {
        'template': template
    })

@login_required(login_url='/admin/login/')
def generate_report(request, template_id):
    template = get_object_or_404(ReportTemplate, id=template_id, user=request.user)
    
    if request.method == 'POST':
        form = GenerateReportForm(request.POST)
        if form.is_valid():
            # Create a new report record
            filter_criteria = {
                'start_date': form.cleaned_data['start_date'].isoformat() if form.cleaned_data['start_date'] else None,
                'end_date': form.cleaned_data['end_date'].isoformat() if form.cleaned_data['end_date'] else None,
                'publication_id': form.cleaned_data['publication'].id if form.cleaned_data['publication'] else None,
                'edition_id': form.cleaned_data['edition'].id if form.cleaned_data['edition'] else None,
                'client_id': form.cleaned_data['client'].id if form.cleaned_data['client'] else None,
                'campaign_id': form.cleaned_data['campaign'].id if form.cleaned_data['campaign'] else None,
                'coverage_type': form.cleaned_data['coverage_type']
            }
            
            report = GeneratedReport.objects.create(
                template=template,
                user=request.user,
                filter_criteria=filter_criteria,
                status='processing'
            )
            
            # Process the report (in a real app, this would be a background task)
            try:
                process_report(report.id)
                messages.success(request, "Report generated successfully.")
                return redirect('download_generated_report', report_id=report.id)
            except Exception as e:
                report.status = 'failed'
                report.error_message = str(e)
                report.save()
                messages.error(request, f"Error generating report: {str(e)}")
    else:
        form = GenerateReportForm()
    
    return render(request, 'core/generate_report.html', {
        'template': template,
        'form': form
    })

@login_required(login_url='/admin/login/')
def download_generated_report(request, report_id):
    report = get_object_or_404(GeneratedReport, id=report_id, user=request.user)
    
    if report.status == 'completed':
        # Serve the file for download
        response = FileResponse(report.report_file, as_attachment=True)
        return response
    
    return render(request, 'core/download_report.html', {
        'report': report
    })

def process_report(report_id):
    """Process and generate the report based on the template and filter criteria."""
    report = GeneratedReport.objects.get(id=report_id)
    template = report.template
    
    try:
        # Get filter criteria
        filter_criteria = report.filter_criteria
        
        # Build query for coverages based on filter criteria
        query = Coverage.objects.all()
        
        if filter_criteria.get('start_date'):
            query = query.filter(date__gte=filter_criteria['start_date'])
        
        if filter_criteria.get('end_date'):
            query = query.filter(date__lte=filter_criteria['end_date'])
        
        if filter_criteria.get('publication_id'):
            query = query.filter(publication_id=filter_criteria['publication_id'])
        
        if filter_criteria.get('edition_id'):
            query = query.filter(edition_id=filter_criteria['edition_id'])
        
        if filter_criteria.get('client_id'):
            query = query.filter(client_id=filter_criteria['client_id'])
        
        if filter_criteria.get('campaign_id'):
            query = query.filter(campaign_id=filter_criteria['campaign_id'])
        
        if filter_criteria.get('coverage_type'):
            query = query.filter(type=filter_criteria['coverage_type'])
        
        # Select related objects to optimize queries
        coverages = query.select_related('publication', 'edition', 'client', 'campaign')
        
        # Load the template Excel file
        from openpyxl.utils import get_column_letter
        import os
        
        # Load template workbook
        wb = openpyxl.load_workbook(template.template_file.path)
        sheet = wb.active
        
        # Find the data start row (assuming headers are in row 1)
        data_start_row = 2
        
        # Get field mappings
        field_mappings = template.field_mappings
        
        # Map Excel columns to their letter references
        column_mapping = {}
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header in field_mappings:
                column_mapping[field_mappings[header]] = col
        
        # Fill in the data
        current_row = data_start_row
        for coverage in coverages:
            for db_field, col in column_mapping.items():
                # Get the value based on the field path
                if '.' in db_field or '__' in db_field:
                    # Handle related fields
                    parts = db_field.replace('__', '.').split('.')
                    value = coverage
                    for part in parts:
                        if hasattr(value, part):
                            value = getattr(value, part)
                        else:
                            value = None
                            break
                else:
                    # Handle direct fields
                    value = getattr(coverage, db_field, None)
                
                # Set the cell value
                sheet.cell(row=current_row, column=col).value = value
            
            current_row += 1
        
        # Save the workbook to a temporary file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"report_{template.name.replace(' ', '_')}_{timestamp}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, 'generated_reports', filename)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Save the workbook
        wb.save(filepath)
        
        # Update the report record
        from django.core.files.base import File
        with open(filepath, 'rb') as f:
            report.report_file.save(filename, File(f), save=False)
        
        report.status = 'completed'
        report.save()
        
        # Clean up the temporary file
        os.remove(filepath)
        
    except Exception as e:
        report.status = 'failed'
        report.error_message = str(e)
        report.save()
        raise e

@login_required(login_url='/admin/login/')
def get_editions(request):
    """AJAX view to get editions for a publication"""
    publication_id = request.GET.get('publication_id')
    if publication_id:
        editions = Edition.objects.filter(publication_id=publication_id).values('id', 'name')
        return JsonResponse({'editions': list(editions)})
    return JsonResponse({'editions': []})

@login_required(login_url='/admin/login/')
def get_campaigns(request):
    """AJAX view to get campaigns for a client"""
    client_id = request.GET.get('client_id')
    if client_id:
        campaigns = Campaign.objects.filter(client_id=client_id).values('id', 'name')
        return JsonResponse({'campaigns': list(campaigns)})
    return JsonResponse({'campaigns': []})

@login_required(login_url='/admin/login/')
def add_publication(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            publication, created = Publication.objects.get_or_create(name=name)
            if created:
                messages.success(request, f"Publication '{name}' created successfully.")
            else:
                messages.info(request, f"Publication '{name}' already exists.")
            
            # Redirect back to search with the new publication selected
            return redirect(f"{reverse('search')}?publication={publication.id}")
        else:
            messages.error(request, "Publication name cannot be empty.")
            return redirect('search')
    
    # GET request - show form
    return render(request, 'core/add_publication.html')

@login_required(login_url='/admin/login/')
def add_edition(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        publication_id = request.POST.get('publication')
        
        if name and publication_id:
            try:
                publication = Publication.objects.get(id=publication_id)
                edition, created = Edition.objects.get_or_create(
                    name=name,
                    publication=publication
                )
                if created:
                    messages.success(request, f"Edition '{name}' for '{publication.name}' created successfully.")
                else:
                    messages.info(request, f"Edition '{name}' for '{publication.name}' already exists.")
                
                # Redirect back to search with the new edition selected
                return redirect(f"{reverse('search')}?edition={edition.id}")
            except Publication.DoesNotExist:
                messages.error(request, "Selected publication does not exist.")
        else:
            messages.error(request, "Edition name and publication are required.")
        
        return redirect('search')
    
    # GET request - show form
    publications = Publication.objects.all().order_by('name')
    return render(request, 'core/add_edition.html', {'publications': publications})

@login_required(login_url='/admin/login/')
def add_client(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            client, created = Client.objects.get_or_create(name=name)
            if created:
                messages.success(request, f"Client '{name}' created successfully.")
            else:
                messages.info(request, f"Client '{name}' already exists.")
            
            # Redirect to add coverage page with the new client selected
            return redirect(f"{reverse('add_coverage')}?client={client.id}")
        else:
            messages.error(request, "Client name cannot be empty.")
            return redirect('search')
    
    # GET request - show form
    return render(request, 'core/add_client.html')

@login_required(login_url='/admin/login/')
def add_campaign(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        client_id = request.POST.get('client')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        if name and client_id and start_date and end_date:
            try:
                client = Client.objects.get(id=client_id)
                campaign = Campaign.objects.create(
                    name=name,
                    client=client,
                    start_date=start_date,
                    end_date=end_date
                )
                messages.success(request, f"Campaign '{name}' for '{client.name}' created successfully.")
                
                # Redirect to add coverage page with the new campaign selected
                return redirect(f"{reverse('add_coverage')}?client={client.id}&campaign={campaign.id}")
            except Client.DoesNotExist:
                messages.error(request, "Selected client does not exist.")
        else:
            messages.error(request, "All fields are required.")
        
        return redirect('search')
    
    # GET request - show form
    client_id = request.GET.get('client', '')
    clients = Client.objects.all().order_by('name')
    selected_client = None
    
    if client_id:
        try:
            selected_client = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            pass
    
    return render(request, 'core/add_campaign.html', {
        'clients': clients,
        'selected_client': selected_client
    })

